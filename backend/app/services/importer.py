"""출장대장 XLSX 일괄 업로드 — exporter.py 와 동일 헤더 형식 지원.

지원 헤더 (한글, exporter 의 COLUMNS 동일):
  No, 제목, 출장자, 부서, 출장일, 일수, 출장지, 거리(km),
  유형, 유류비, 톨게이트, 대중교통, 식비, 일비, 숙박비, 합계,
  예산명, 시스템, 자금집행일, 상태

- 헤더는 첫 행으로 가정 (다른 위치면 첫 행을 헤더로 인식)
- "합계" 행(No 셀이 "합계") 자동 스킵
- 누락된 컬럼은 기본값으로 채움 — 필수: 출장자, 출장일, 출장지, 제목
- 유형(모드)은 한글 → 영문 자동 매핑
"""

from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

MODE_REVERSE = {
    "자가운전": "self_drive",
    "자가차량 운전": "self_drive",
    "자가동승": "self_passenger",
    "자가차량 동승": "self_passenger",
    "공용차량": "company_car",
    "대중교통": "public_transit",
    # 영문 그대로도 허용
    "self_drive": "self_drive",
    "self_passenger": "self_passenger",
    "company_car": "company_car",
    "public_transit": "public_transit",
}

# 헤더 한글 → Trip 필드명 매핑
HEADER_MAP = {
    "No": "no",
    "no": "no",
    "제목": "title",
    "출장자": "traveler_name",
    "부서": "dept",
    "출장일": "trip_date",
    "일수": "days",
    "출장지": "place",
    "거리(km)": "distance_km",
    "거리": "distance_km",
    "유형": "mode",
    "모드": "mode",
    "유류비": "fuel_cost",
    "톨게이트": "toll_sum",
    "대중교통": "public_cost",
    "식비": "meal_cost",
    "일비": "per_diem",
    "숙박비": "lodge_cost",
    "합계": "total",
    "예산명": "biz_name",
    "사업명": "biz_name",
    "시스템": "fund_system",
    "자금집행일": "fund_date",
    "상태": "status",
}


class ImportError(ValueError):
    pass


def _to_date(v) -> date | None:
    if v is None or v == "":
        return None
    if isinstance(v, date) and not isinstance(v, datetime):
        return v
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, str):
        for fmt in ("%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d"):
            try:
                return datetime.strptime(v.strip(), fmt).date()
            except ValueError:
                continue
    return None


def _to_int(v) -> int:
    if v is None or v == "":
        return 0
    if isinstance(v, (int, float)):
        return int(v)
    if isinstance(v, str):
        s = v.replace(",", "").replace("원", "").strip()
        try:
            return int(float(s)) if s else 0
        except ValueError:
            return 0
    return 0


def _to_float(v) -> float | None:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.replace(",", "").replace("km", "").strip()
        try:
            return float(s) if s else None
        except ValueError:
            return None
    return None


def _to_mode(v) -> str:
    if not v:
        return "self_drive"
    s = str(v).strip()
    return MODE_REVERSE.get(s, "self_drive")


def parse_xlsx(content: bytes) -> tuple[list[dict[str, Any]], list[str]]:
    """XLSX 바이트를 파싱해 trips dict 리스트 + 경고 리스트 반환.

    반환 dict 은 backend 의 Trip 모델 필드명 (영문) 으로 정규화.
    """
    try:
        wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        raise ImportError(f"XLSX 파일을 열 수 없습니다: {e}")

    ws = wb.active
    if ws.max_row < 2:
        raise ImportError("데이터 행이 없습니다 (헤더 1행 + 데이터 행 필요)")

    # 헤더 매핑
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    col_to_field: dict[int, str] = {}
    for i, h in enumerate(header_row):
        if h is None:
            continue
        key = str(h).strip()
        if key in HEADER_MAP:
            col_to_field[i] = HEADER_MAP[key]

    required = {"traveler_name", "trip_date", "place", "title"}
    have = set(col_to_field.values())
    missing = required - have
    if missing:
        # 가장 흔한 누락은 한글 헤더 오타 — 경고로 어떤 컬럼이 필요한지 알려줌
        raise ImportError(
            f"필수 컬럼 누락: {', '.join(sorted(missing))}. "
            f"엑셀 헤더는 다음 중 하나여야 합니다 — "
            f"출장자/출장일/출장지/제목"
        )

    trips: list[dict[str, Any]] = []
    warnings: list[str] = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # 합계 행 스킵 (No 셀이 "합계" 또는 모든 셀이 빈 행)
        if all(c is None or c == "" for c in row):
            continue
        no_val = row[0] if len(row) > 0 else None
        if isinstance(no_val, str) and no_val.strip() == "합계":
            continue

        # 컬럼 → 값
        rec: dict[str, Any] = {}
        for col_idx, field in col_to_field.items():
            if col_idx >= len(row):
                continue
            rec[field] = row[col_idx]

        # 정규화 + 검증
        try:
            trip = {
                "no": int(rec["no"]) if rec.get("no") else None,
                "title": str(rec.get("title") or "").strip(),
                "traveler_name": str(rec.get("traveler_name") or "").strip(),
                "dept": str(rec.get("dept") or "").strip() or None,
                "trip_date": _to_date(rec.get("trip_date")),
                "days": int(rec.get("days") or 1) or 1,
                "place": str(rec.get("place") or "").strip(),
                "distance_km": _to_float(rec.get("distance_km")),
                "mode": _to_mode(rec.get("mode")),
                "fuel_cost": _to_int(rec.get("fuel_cost")),
                "toll_sum": _to_int(rec.get("toll_sum")),
                "public_cost": _to_int(rec.get("public_cost")),
                "meal_cost": _to_int(rec.get("meal_cost")),
                "per_diem": _to_int(rec.get("per_diem")),
                "lodge_cost": _to_int(rec.get("lodge_cost")),
                "total": _to_int(rec.get("total")),
                "biz_name": str(rec.get("biz_name") or "").strip() or None,
                "fund_system": str(rec.get("fund_system") or "").strip() or None,
                "fund_date": _to_date(rec.get("fund_date")),
                "status": str(rec.get("status") or "확정").strip() or "확정",
            }
        except (ValueError, TypeError) as e:
            warnings.append(f"{row_idx}행 파싱 실패 — 건너뜀: {e}")
            continue

        # 필수 검증
        if not trip["traveler_name"]:
            warnings.append(f"{row_idx}행 출장자 비어있음 — 건너뜀")
            continue
        if not trip["trip_date"]:
            warnings.append(f"{row_idx}행 출장일 비어있음 — 건너뜀")
            continue
        if not trip["place"]:
            warnings.append(f"{row_idx}행 출장지 비어있음 — 건너뜀")
            continue
        if not trip["title"]:
            # 제목 누락이면 자동 생성
            md = trip["trip_date"]
            trip["title"] = f"({md.month:02d}.{md.day:02d}/{trip['place']}) 출장"

        # total 누락이면 자동 합산
        if trip["total"] == 0:
            trip["total"] = (
                trip["fuel_cost"] + trip["toll_sum"] + trip["public_cost"]
                + trip["meal_cost"] + trip["per_diem"] + trip["lodge_cost"]
            )

        trips.append(trip)

    return trips, warnings
