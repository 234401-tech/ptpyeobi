"""출장대장 XLSX export — openpyxl.

엑셀 `2026 국내②` 시트 형식에 가깝게 구성:
  No / 제목 / 출장자 / 부서 / 출장일 / 일수 / 출장지 / 거리(km) /
  유형 / 유류비 / 톨게이트 / 대중교통 / 식비 / 일비 / 숙박비 / 합계 /
  예산명 / 시스템 / 자금집행일 / 상태
"""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models import Trip

MODE_LABEL = {
    "self_drive": "자가운전",
    "self_passenger": "자가동승",
    "company_car": "공용차량",
    "public_transit": "대중교통",
}

COLUMNS = [
    ("No", 6),
    ("제목", 36),
    ("출장자", 8),
    ("부서", 14),
    ("출장일", 12),
    ("일수", 5),
    ("출장지", 18),
    ("거리(km)", 8),
    ("유형", 8),
    ("유류비", 10),
    ("톨게이트", 10),
    ("대중교통", 10),
    ("식비", 10),
    ("일비", 10),
    ("숙박비", 10),
    ("합계", 12),
    ("예산명", 22),
    ("시스템", 8),
    ("자금집행일", 12),
    ("상태", 8),
]

HEADER_FILL = PatternFill("solid", fgColor="4338CA")  # indigo-700
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
THIN = Side(style="thin", color="E2E8F0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)
RIGHT = Alignment(horizontal="right", vertical="center")
MONEY_FMT = "#,##0"


def trips_to_xlsx(trips: list[Trip]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "2026 국내②"

    # 헤더
    for i, (name, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=i, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"

    # 데이터
    for r, t in enumerate(trips, start=2):
        row_values = [
            t.no,
            t.title,
            t.traveler_name,
            t.dept,
            t.trip_date,
            t.days,
            t.place,
            t.distance_km,
            MODE_LABEL.get(t.mode, t.mode),
            t.fuel_cost,
            t.toll_sum,
            t.public_cost,
            t.meal_cost,
            t.per_diem,
            t.lodge_cost,
            t.total,
            t.biz_name,
            t.fund_system,
            t.fund_date,
            t.status,
        ]
        for c, v in enumerate(row_values, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = BORDER
            # 컬럼별 정렬/포맷
            if c in (1, 3, 6, 8, 9, 18, 20):
                cell.alignment = CENTER
            elif c in (10, 11, 12, 13, 14, 15, 16):
                cell.alignment = RIGHT
                cell.number_format = MONEY_FMT
            elif c in (5, 19):
                cell.alignment = CENTER
                cell.number_format = "yyyy-mm-dd"
            else:
                cell.alignment = LEFT

    # 합계 행
    if trips:
        sum_row = len(trips) + 2
        ws.cell(row=sum_row, column=1, value="합계").alignment = CENTER
        ws.cell(row=sum_row, column=1).font = Font(bold=True)
        for c in (10, 11, 12, 13, 14, 15, 16):
            col_letter = get_column_letter(c)
            cell = ws.cell(
                row=sum_row,
                column=c,
                value=f"=SUM({col_letter}2:{col_letter}{sum_row - 1})",
            )
            cell.number_format = MONEY_FMT
            cell.font = Font(bold=True)
            cell.alignment = RIGHT
            cell.border = BORDER

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
